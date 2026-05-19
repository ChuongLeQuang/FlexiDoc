"""
EN: Core service for processing and extracting data from Excel files.
VI: Dịch vụ lõi để xử lý và trích xuất dữ liệu từ tệp Excel.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from src.exceptions.custom_errors import (
    SheetNotFoundError,
    EmptyHeaderRowError,
    InvalidFileFormatError,
)


def validate_sheet(excel_bytes: bytes, sheet_name: str) -> bool:
    """
    EN: Check if the specified sheet exists in the workbook.
    VI: Kiểm tra xem sheet được chỉ định có tồn tại trong tệp không.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes), read_only=True, data_only=True
        )
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(
                f"Không tìm thấy sheet '{sheet_name}' trong file Excel."
            )
        return True
    except SheetNotFoundError:
        raise
    except Exception:
        raise InvalidFileFormatError("Định dạng file Excel không hợp lệ hoặc bị hỏng.")
    finally:
        if "wb" in locals():
            wb.close()


def normalize_headers(
    sheet_obj: openpyxl.worksheet.worksheet.Worksheet, header_row: int
) -> list[str]:
    """
    EN: Extract and normalize header strings.
    VI: Trích xuất và chuẩn hóa dòng tiêu đề.
    """
    headers = []
    for row in sheet_obj.iter_rows(
        min_row=header_row, max_row=header_row, values_only=True
    ):
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                val = str(cell).strip()
                # Loại bỏ khoảng trắng kép ở giữa chữ
                val = " ".join(val.split())
                headers.append(val)
            else:
                headers.append("")
        break  # Chỉ đọc dòng header

    if not any(headers):
        raise EmptyHeaderRowError(f"Dòng tiêu đề (dòng {header_row}) hoàn toàn trống.")
    return headers


def extract_raw_data(
    excel_bytes: bytes, sheet_name: str, header_row: int = 1
) -> list[dict]:
    """
    EN: Extract raw data, handling merged cells and skipping empty rows.
    VI: Trích xuất dữ liệu thô, xử lý tự động điền ô gộp và bỏ qua dòng trống.
    """
    validate_sheet(excel_bytes, sheet_name)  # Đảm bảo file hợp lệ trước

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb[sheet_name]

    # Xử lý ô gộp: Unmerge và copy giá trị từ ô góc trái trên cùng xuống các ô còn lại
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_val

    headers = normalize_headers(ws, header_row)
    data = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue  # Bỏ qua dòng trống tinh

        row_dict = {
            headers[idx]: str(cell).strip() if cell is not None else ""
            for idx, cell in enumerate(row)
            if idx < len(headers) and headers[idx]
        }
        data.append(row_dict)

    wb.close()
    return data


def create_empty_template(variables: list[str]) -> bytes:
    """
    EN: Create an empty Excel template with variables as headers.
    VI: Tạo một file Excel mẫu trống với các biến làm tiêu đề.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Tạo tiêu đề và bôi màu xám, in đậm cho đẹp
    header_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    header_font = Font(bold=True)

    for col_num, var_name in enumerate(variables, 1):
        cell = ws.cell(row=1, column=col_num, value=var_name)
        cell.font = header_font
        cell.fill = header_fill

        # Tạo 1 dòng dữ liệu ví dụ ở ngay bên dưới
        ws.cell(row=2, column=col_num, value=f"Ví dụ: {var_name}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
