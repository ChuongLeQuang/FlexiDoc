"""
EN: Unit tests for Assistant Module (Tag Builder).
VI: Kịch bản kiểm thử cho Module Trợ lý Sinh Thẻ.
"""

import pytest
import io
import openpyxl
from fastapi.testclient import TestClient
from src.main import app

from src.services.excel_service import create_empty_template

client = TestClient(app)


def test_create_empty_template():
    """Kiểm tra hàm sinh file Excel mẫu trên RAM."""
    variables = ["ho_ten", "ngay_sinh", "muc_luong"]
    excel_bytes = create_empty_template(variables)

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    # Kiểm tra dòng Header
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(variables) + 1)]
    assert headers == variables

    # Kiểm tra xem có dữ liệu ví dụ ở dòng 2 không
    dummy_data = ws.cell(row=2, column=1).value
    assert dummy_data is not None


def test_assistant_api_generate_excel():
    """Kiểm tra API sinh file Excel mẫu."""
    variables = ["ho_ten", "item.ten_lop"]
    response = client.post("/api/v1/assistant/generate-excel", json=variables)

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
