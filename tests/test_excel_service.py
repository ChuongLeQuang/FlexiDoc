"""
EN: Unit tests for Excel Service.
VI: Kịch bản kiểm thử cho Dịch vụ Excel.
"""

import pytest
import io
import openpyxl
from src.services.excel_service import (
    validate_sheet,
    normalize_headers,
    extract_raw_data,
)
from src.exceptions.custom_errors import SheetNotFoundError, EmptyHeaderRowError


def create_dummy_excel(sheet_name="Sheet1", data=None, merged_cells=None) -> bytes:
    """Tạo file Excel ảo trên RAM."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        for r_idx, row in enumerate(data, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    if merged_cells:
        for merge_range in merged_cells:
            ws.merge_cells(merge_range)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_validate_sheet_not_found():
    """Test lỗi khi không tìm thấy sheet."""
    excel_bytes = create_dummy_excel(sheet_name="DataSheet")
    with pytest.raises(SheetNotFoundError):
        validate_sheet(excel_bytes, "WrongSheet")


def test_normalize_headers():
    """Test chuẩn hóa dòng tiêu đề."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([" Họ và tên  ", "Lương   mới", None])

    headers = normalize_headers(ws, 1)
    assert headers == ["Họ và tên", "Lương mới", ""]


def test_normalize_headers_empty():
    """Test lỗi dòng tiêu đề rỗng."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, "   ", ""])

    with pytest.raises(EmptyHeaderRowError):
        normalize_headers(ws, 1)


def test_extract_raw_data_merged():
    """Test trích xuất và unmerge tự động điền dữ liệu (forward-fill)."""
    data = [
        ["Phòng ban", "Nhân viên"],
        ["IT", "Alice"],
        [None, "Bob"],  # Chỗ None này thuộc vùng bị gộp A2:A3 với chữ "IT"
    ]
    excel_bytes = create_dummy_excel(data=data, merged_cells=["A2:A3"])
    result = extract_raw_data(excel_bytes, "Sheet1", 1)

    assert len(result) == 2
    assert result[0]["Phòng ban"] == "IT"
    assert result[1]["Phòng ban"] == "IT"  # Phải được tự động điền
    assert result[1]["Nhân viên"] == "Bob"


def test_extract_raw_data_empty_rows():
    """Test bỏ qua dòng trống tinh."""
    data = [["Họ tên"], ["Alice"], [None], ["   "]]
    excel_bytes = create_dummy_excel(data=data)
    result = extract_raw_data(excel_bytes, "Sheet1", 1)

    assert len(result) == 1
    assert result[0]["Họ tên"] == "Alice"
