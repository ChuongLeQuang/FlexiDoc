"""
EN: UI Component for triggering document generation and downloading results.
VI: Thành phần giao diện kích hoạt sinh tài liệu và tải kết quả.
"""

import streamlit as st
import time
import io
from typing import Dict, Any, List
import frontend.api_client as api_client


def get_unique_values(
    excel_bytes: bytes, sheet_name: str, header_row: int, col_name: str
) -> List[str]:
    """Trích xuất danh sách các giá trị duy nhất từ một cột trong Excel để làm bộ lọc."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes), read_only=True, data_only=True
        )
        ws = wb[sheet_name]
        headers = []
        for row in ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True
        ):
            headers = [str(cell).strip() if cell else "" for cell in row]
            break

        if col_name not in headers:
            wb.close()
            return []

        col_idx = headers.index(col_name)
        values = set()
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            val = row[col_idx]
            if val is not None and str(val).strip() != "":
                values.add(str(val).strip())
        wb.close()
        return sorted(list(values))
    except Exception:
        return []


def handle_generation(
    mapping_config: Dict[str, Any], upload_data: Dict[str, Any]
) -> None:
    """
    EN: Render generation button and download link.
    VI: Hiển thị nút tạo tài liệu và link tải về.
    """
    st.subheader("🎯 Tùy chọn Lọc Dữ liệu")

    excel_cols = upload_data.get("excel_cols", [])

    # Tìm index mặc định (Ưu tiên Cột khóa gom nhóm hoặc cột map đầu tiên để UX tốt hơn)
    default_filter_col = mapping_config.get("grouping_key")
    default_idx = 0
    if default_filter_col in excel_cols:
        default_idx = excel_cols.index(default_filter_col)
    elif (
        mapping_config.get("rules")
        and mapping_config["rules"][0].get("excel_col") in excel_cols
    ):
        default_idx = excel_cols.index(mapping_config["rules"][0].get("excel_col"))

    filter_col = ""
    if excel_cols:
        filter_col = st.selectbox(
            "Chọn cột làm tiêu chí lọc:", options=excel_cols, index=default_idx
        )

    mapping_config["filter_col"] = filter_col

    unique_vals = []
    if filter_col:
        unique_vals = get_unique_values(
            upload_data["xlsx_bytes"],
            upload_data["sheet_name"],
            upload_data["header_row"],
            filter_col,
        )

    selected_keys = []
    if unique_vals:
        selected_keys = st.multiselect(
            f"Chọn các giá trị từ cột '{filter_col}' (Để trống để tạo cho TẤT CẢ):",
            options=unique_vals,
            help="Hệ thống sẽ chỉ tạo file Word cho những dòng dữ liệu thỏa mãn điều kiện này.",
        )
    mapping_config["selected_keys"] = selected_keys

    st.divider()
    st.write(
        "Sau khi hoàn tất cấu hình, bấm nút bên dưới để tiến hành trộn dữ liệu và xuất file ZIP."
    )

    if st.button("🚀 Bắt đầu Tạo Tài liệu", type="primary", use_container_width=True):
        with st.spinner("Đang xử lý dữ liệu và tạo file Word..."):
            payload = {"mapping_config": mapping_config, "upload_data": upload_data}
            try:
                zip_bytes = api_client.generate_document(payload)
                st.success(
                    "✅ Đã tạo tài liệu thành công! Vui lòng tải file ZIP về máy."
                )

                st.download_button(
                    label="📥 Tải xuống kết quả (.zip)",
                    data=zip_bytes,
                    file_name="FlexiDoc_Results.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(str(e))
